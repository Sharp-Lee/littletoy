#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
转粮数据汇总脚本
读取所有转粮数据文件中的"总产出重量"，按顺序汇总到汇总.xlsx文件
"""

import pandas as pd
import os
from pathlib import Path
import openpyxl
import re

def is_valid_weight(value):
    """
    判断一个值是否是有效的重量值
    
    Args:
        value: 要检查的值
        
    Returns:
        如果是有效的重量值则返回浮点数，否则返回None
    """
    if value is None:
        return None
    
    try:
        weight = float(value)
        # 排除明显不是重量的值：负数、0、过小的值（可能是百分比）、过大的值（可能是日期时间戳）
        if weight > 0 and weight < 1e10:  # 合理的重量范围：0到100亿
            # 排除日期时间格式（通常是很大的数字）
            if weight > 1000000000:  # 可能是时间戳
                return None
            return weight
    except (ValueError, TypeError):
        pass
    
    return None

def find_value_near_label(ws, label_row, label_col, max_search_range=15):
    """
    在标签附近查找数值（多种策略）
    
    Args:
        ws: 工作表对象
        label_row: 标签所在行
        label_col: 标签所在列
        max_search_range: 最大搜索范围
        
    Returns:
        找到的数值，如果找不到则返回None
    """
    candidates = []
    
    # 策略1: 同一行右侧查找（最常见的情况）
    for offset in range(1, min(max_search_range + 1, ws.max_column - label_col + 1)):
        target_col = label_col + offset
        if target_col <= ws.max_column:
            cell = ws.cell(row=label_row, column=target_col)
            weight = is_valid_weight(cell.value)
            if weight:
                candidates.append((weight, f"同一行右侧{offset}列"))
    
    # 策略2: 同一行左侧查找（以防标签在数值右侧）
    for offset in range(1, min(max_search_range + 1, label_col)):
        target_col = label_col - offset
        if target_col >= 1:
            cell = ws.cell(row=label_row, column=target_col)
            weight = is_valid_weight(cell.value)
            if weight:
                candidates.append((weight, f"同一行左侧{offset}列"))
    
    # 策略3: 下一行查找（数值可能在标签下方）
    if label_row < ws.max_row:
        next_row = label_row + 1
        # 在下一行的附近列查找
        for offset in range(-2, min(max_search_range + 1, ws.max_column - label_col + 3)):
            target_col = label_col + offset
            if 1 <= target_col <= ws.max_column:
                cell = ws.cell(row=next_row, column=target_col)
                weight = is_valid_weight(cell.value)
                if weight:
                    candidates.append((weight, f"下一行第{target_col}列"))
    
    # 策略4: 同一列下方查找
    for offset in range(1, min(max_search_range + 1, ws.max_row - label_row + 1)):
        target_row = label_row + offset
        if target_row <= ws.max_row:
            cell = ws.cell(row=target_row, column=label_col)
            weight = is_valid_weight(cell.value)
            if weight:
                candidates.append((weight, f"同一列下方{offset}行"))
    
    # 策略5: 下一行的同一列位置查找（最常见的备选位置）
    if label_row < ws.max_row:
        cell = ws.cell(row=label_row + 1, column=label_col)
        weight = is_valid_weight(cell.value)
        if weight:
            candidates.append((weight, "下一行同一列"))
    
    # 返回第一个找到的有效值（按优先级：右侧 > 下方 > 左侧）
    if candidates:
        # 优先选择右侧的值
        right_candidates = [c for c in candidates if "右侧" in c[1]]
        if right_candidates:
            return right_candidates[0][0]
        # 其次选择下方的值
        down_candidates = [c for c in candidates if "下方" in c[1] or "下一行" in c[1]]
        if down_candidates:
            return down_candidates[0][0]
        # 最后选择左侧的值
        return candidates[0][0]
    
    return None

def find_total_output_weight(file_path):
    """
    从转粮数据文件中查找"总产出重量"的值
    使用多种策略确保在格式变化时仍能准确找到
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        总产出重量的值，如果找不到则返回None
    """
    try:
        # 根据文件扩展名选择不同的读取方式
        is_xlsx = file_path.endswith('.xlsx')
        
        if is_xlsx:
            # 使用openpyxl读取.xlsx文件
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # 遍历所有工作表
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 遍历所有行查找"总产出重量"
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=False), 1):
                    for col_idx, cell in enumerate(row, 1):
                        cell_value = str(cell.value).strip() if cell.value else ""
                        
                        # 检查是否包含"总产出重量"关键词（支持部分匹配）
                        if "总产出重量" in cell_value or ("总产出" in cell_value and "重量" in cell_value):
                            # 使用多种策略在附近查找数值
                            weight = find_value_near_label(ws, row_idx, col_idx)
                            if weight:
                                return weight
        else:
            # 使用pandas读取.xls文件（使用xlrd引擎）
            excel_file = pd.ExcelFile(file_path, engine='xlrd')
            
            # 遍历所有工作表
            for sheet_name in excel_file.sheet_names:
                # 读取整个工作表
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd', header=None)
                
                # 遍历所有行查找"总产出重量"
                for row_idx in range(len(df)):
                    row = df.iloc[row_idx]
                    
                    # 遍历该行的所有列
                    for col_idx in range(len(row)):
                        cell_value = str(row.iloc[col_idx]).strip() if pd.notna(row.iloc[col_idx]) else ""
                        
                        # 检查是否包含"总产出重量"关键词
                        if "总产出重量" in cell_value or ("总产出" in cell_value and "重量" in cell_value):
                            # 策略1: 同一行右侧查找
                            for offset in range(1, min(16, len(row) - col_idx)):
                                target_col_idx = col_idx + offset
                                if target_col_idx < len(row):
                                    weight = is_valid_weight(row.iloc[target_col_idx])
                                    if weight:
                                        return weight
                            
                            # 策略2: 下一行查找
                            if row_idx + 1 < len(df):
                                next_row = df.iloc[row_idx + 1]
                                # 在下一行的附近列查找
                                for offset in range(-2, min(16, len(next_row) - col_idx + 2)):
                                    target_col_idx = col_idx + offset
                                    if 0 <= target_col_idx < len(next_row):
                                        weight = is_valid_weight(next_row.iloc[target_col_idx])
                                        if weight:
                                            return weight
                            
                            # 策略3: 同一列下方查找
                            for offset in range(1, min(16, len(df) - row_idx)):
                                target_row_idx = row_idx + offset
                                if target_row_idx < len(df):
                                    weight = is_valid_weight(df.iloc[target_row_idx, col_idx])
                                    if weight:
                                        return weight
        
        return None
    except Exception as e:
        print(f"读取文件 {file_path} 时出错: {e}")
        return None

def extract_file_number(file_name):
    """
    从文件名中提取文件编号（支持任意长度的数字）
    
    Args:
        file_name: 文件名，如"709转粮数据.xls"或"909转粮数据.xls"
        
    Returns:
        文件编号字符串，如果提取失败则返回None
    """
    # 使用正则表达式提取文件名开头的数字
    match = re.match(r'^(\d+)', file_name)
    if match:
        return match.group(1)
    return None

def create_summary_file(summary_file, data_dict):
    """
    创建新的汇总文件
    
    Args:
        summary_file: 汇总文件路径
        data_dict: 字典，键为文件编号（如'709'），值为总产出重量
    """
    try:
        # 创建新的工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 按文件编号的数字大小排序
        sorted_file_nums = sorted(data_dict.keys(), key=lambda x: int(x))
        
        # 第一行：表头
        # A1: 中心名词
        ws.cell(row=1, column=1, value="中心名词")
        
        # B1, C1, ...: 文件编号
        for col_idx, file_num in enumerate(sorted_file_nums, start=2):
            ws.cell(row=1, column=col_idx, value=file_num)
        
        # 第二行：数量行
        # A2: 数量
        ws.cell(row=2, column=1, value="数量")
        
        # B2, C2, ...: 总产出重量值
        for col_idx, file_num in enumerate(sorted_file_nums, start=2):
            weight = data_dict[file_num]
            ws.cell(row=2, column=col_idx, value=weight)
            print(f"  ✓ 已写入 {file_num} 列（第2行，第{col_idx}列）: {weight}")
        
        # 保存文件
        wb.save(summary_file)
        print(f"\n✓ 已创建新汇总文件: {summary_file}")
        print(f"  共 {len(sorted_file_nums)} 个文件编号列")
        
    except Exception as e:
        print(f"创建汇总文件时出错: {e}")
        import traceback
        traceback.print_exc()
        raise

def update_summary_file(summary_file, data_dict):
    """
    更新汇总文件，将总产出重量写入"数量"行对应的列
    支持动态文件编号范围，如果汇总文件中没有对应的列，会自动添加
    
    汇总文件格式：
    第一行：中心名词 | 709 | 710 | 711 | ... | (动态编号)
    第二行：数量 | (总产出重量值) | (总产出重量值) | ... | (总产出重量值)
    
    Args:
        summary_file: 汇总文件路径
        data_dict: 字典，键为文件编号（如'709'），值为总产出重量
    """
    try:
        # 使用openpyxl打开工作簿
        wb = openpyxl.load_workbook(summary_file)
        ws = wb.active
        
        # 查找表头行（第一行，包含文件编号）
        header_row = 1
        column_mapping = {}  # 文件编号 -> 列索引
        existing_numbers = set()  # 已存在的文件编号
        
        # 从第一行查找所有已存在的文件编号
        for col_idx, cell in enumerate(ws.iter_cols(min_col=1, max_col=ws.max_column, 
                                                     min_row=header_row, max_row=header_row, 
                                                     values_only=False), 1):
            cell_value = str(cell[0].value).strip() if cell[0].value else ""
            # 检查是否是文件编号（支持任意长度的数字）
            if cell_value.isdigit():
                file_num = cell_value
                existing_numbers.add(file_num)
                if file_num in data_dict:
                    column_mapping[file_num] = col_idx
                    print(f"  找到已存在的列 {file_num} 在列索引 {col_idx}")
        
        # 查找"数量"行（遍历所有行查找，优先选择最下方的）
        quantity_row = None
        
        # 从下往上查找"数量"标签（通常最下方的才是实际数据行）
        for row_idx in range(ws.max_row, 0, -1):
            cell_value = str(ws.cell(row=row_idx, column=1).value).strip() if ws.cell(row=row_idx, column=1).value else ""
            if "数量" in cell_value:
                quantity_row = row_idx
                break
        
        # 如果找不到"数量"行，报错
        if quantity_row is None:
            raise ValueError("未找到'数量'行，请检查汇总文件格式")
        
        print(f"  找到'数量'行在第 {quantity_row} 行")
        
        # 按文件编号的数字大小排序（而不是字符串排序）
        sorted_file_nums = sorted(data_dict.keys(), key=lambda x: int(x))
        
        # 查找最后一列的位置（用于添加新列）
        last_data_col = 1
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value is not None and str(cell_value).strip():
                last_data_col = col_idx
        
        # 写入数据到"数量"行对应的列
        written_count = 0
        new_columns_added = 0
        
        for file_num in sorted_file_nums:
            weight = data_dict[file_num]
            
            if file_num in column_mapping:
                # 列已存在，直接写入
                col_idx = column_mapping[file_num]
                ws.cell(row=quantity_row, column=col_idx, value=weight)
                print(f"  ✓ 已写入 {file_num} 列（第{quantity_row}行，第{col_idx}列）: {weight}")
                written_count += 1
            else:
                # 列不存在，需要添加新列
                new_col_idx = last_data_col + 1 + new_columns_added
                
                # 在表头行添加文件编号
                ws.cell(row=header_row, column=new_col_idx, value=file_num)
                
                # 在数量行写入数据
                ws.cell(row=quantity_row, column=new_col_idx, value=weight)
                
                print(f"  ✓ 已添加新列 {file_num}（第{header_row}行表头，第{quantity_row}行数据，第{new_col_idx}列）: {weight}")
                written_count += 1
                new_columns_added += 1
        
        # 保存文件
        wb.save(summary_file)
        print(f"\n✓ 汇总文件已更新，共写入 {written_count} 个值")
        if new_columns_added > 0:
            print(f"  （其中新增 {new_columns_added} 个列）")
        
    except Exception as e:
        print(f"更新汇总文件时出错: {e}")
        import traceback
        traceback.print_exc()
        raise

def get_base_dir():
    """
    获取程序运行目录
    兼容Python脚本和PyInstaller打包的exe文件
    """
    if getattr(sys, 'frozen', False):
        # PyInstaller打包的exe文件
        base_dir = Path(sys.executable).parent
    else:
        # Python脚本
        base_dir = Path(__file__).parent
    return base_dir

def main():
    # 获取脚本所在目录（兼容exe打包）
    base_dir = get_base_dir()
    
    # 查找所有转粮数据文件
    all_files = [f for f in os.listdir(base_dir) 
                 if f.endswith(('.xls', '.xlsx')) and '转粮数据' in f]
    
    # 提取文件编号并按数字大小排序（而不是字符串排序）
    files_with_numbers = []
    for file_name in all_files:
        file_number = extract_file_number(file_name)
        if file_number:
            files_with_numbers.append((int(file_number), file_name, file_number))
        else:
            print(f"  警告：无法从文件名 '{file_name}' 中提取文件编号，将跳过")
    
    # 按文件编号的数字大小排序
    files_with_numbers.sort(key=lambda x: x[0])
    data_files = [f[1] for f in files_with_numbers]
    
    print(f"找到 {len(data_files)} 个转粮数据文件")
    if data_files:
        file_numbers = [f[2] for f in files_with_numbers]
        print(f"文件编号范围: {file_numbers[0]} - {file_numbers[-1]}")
    
    # 存储汇总数据：文件编号 -> 总产出重量
    data_dict = {}
    
    # 读取每个文件的总产出重量
    for file_name in data_files:
        file_path = base_dir / file_name
        print(f"正在处理: {file_name}")
        
        weight = find_total_output_weight(str(file_path))
        
        if weight is not None:
            # 提取文件编号（支持任意长度的数字）
            file_number = extract_file_number(file_name)
            if file_number:
                data_dict[file_number] = weight
                print(f"  ✓ 找到总产出重量: {weight}")
            else:
                print(f"  ✗ 无法提取文件编号")
        else:
            print(f"  ✗ 未找到总产出重量")
    
    # 更新或创建汇总文件
    if data_dict:
        summary_file = base_dir / '汇总.xlsx'
        
        if summary_file.exists():
            print(f"\n正在更新汇总文件: {summary_file}")
            update_summary_file(str(summary_file), data_dict)
            print(f"✓ 汇总完成！共处理 {len(data_dict)} 个文件")
        else:
            print(f"\n汇总文件不存在，正在创建新文件: {summary_file}")
            create_summary_file(str(summary_file), data_dict)
            print(f"✓ 汇总完成！共处理 {len(data_dict)} 个文件")
    else:
        print("\n✗ 未找到任何总产出重量数据")

if __name__ == "__main__":
    main()

